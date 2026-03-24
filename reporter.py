from typing import Optional, List, Tuple
"""照合結果をExcelレポートとして出力"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

from matcher import (
    STATUS_OK, STATUS_OUTSIDE_HOURS, STATUS_NO_ATTENDANCE,
    STATUS_PERSON_MISMATCH, STATUS_CONTRACT_ONLY, STATUS_CONTRACT_OUTSIDE,
    STATUS_NO_TIME_INFO, generate_summary
)

# セルの背景色定義
COLOR_OK = "FFFFFF"               # 白（問題なし）
COLOR_OUTSIDE_HOURS = "FFEB9C"    # 黄色（勤務時間外）
COLOR_CONTRACT_OUTSIDE = "FFD966" # 濃い黄色（契約時間外）
COLOR_NO_ATTENDANCE = "FCE4D6"    # オレンジ（出勤記録なし）
COLOR_PERSON_MISMATCH = "FF6B6B"  # 赤（人物不一致）
COLOR_CONTRACT_ONLY = "E2EFDA"    # 薄緑（契約時間参照・問題なし）
COLOR_NO_TIME_INFO = "EDEDED"     # グレー（時刻情報なし）

STATUS_COLOR_MAP = {
    STATUS_OK: COLOR_OK,
    STATUS_OUTSIDE_HOURS: COLOR_OUTSIDE_HOURS,
    STATUS_CONTRACT_OUTSIDE: COLOR_CONTRACT_OUTSIDE,
    STATUS_NO_ATTENDANCE: COLOR_NO_ATTENDANCE,
    STATUS_PERSON_MISMATCH: COLOR_PERSON_MISMATCH,
    STATUS_CONTRACT_ONLY: COLOR_CONTRACT_ONLY,
    STATUS_NO_TIME_INFO: COLOR_NO_TIME_INFO,
}

STATUS_LABEL_MAP = {
    STATUS_OK: "OK",
    STATUS_OUTSIDE_HOURS: "⚠ 勤務時間外",
    STATUS_CONTRACT_OUTSIDE: "⚠ 契約時間外",
    STATUS_NO_ATTENDANCE: "⚠ 出勤記録なし",
    STATUS_PERSON_MISMATCH: "🚨 人物不一致",
    STATUS_CONTRACT_ONLY: "ℹ 契約時間参照",
    STATUS_NO_TIME_INFO: "時刻情報なし",
}


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_time(t) -> str:
    if t is None:
        return ""
    if hasattr(t, "strftime"):
        return t.strftime("%H:%M")
    return str(t)


def _fmt_date(d) -> str:
    if d is None:
        return ""
    if hasattr(d, "strftime"):
        return d.strftime("%Y/%m/%d")
    return str(d)


def generate_report(results: List[dict], output_path: Optional[str] = None) -> bytes:
    """
    照合結果をExcelファイルに出力。
    output_path が None の場合はバイト列で返す。
    """
    wb = openpyxl.Workbook()

    # ============================================================
    # シート1: 照合結果（カラーマーカー付き）
    # ============================================================
    ws_result = wb.active
    ws_result.title = "照合結果"

    headers = [
        "行番号", "社員番号", "氏名", "日付", "視聴開始", "視聴終了",
        "動画タイトル", "進捗率", "ステータス", "判定詳細",
        "出勤打刻（IN）", "出勤打刻（OUT）", "参照元"
    ]

    # ヘッダー行
    header_fill = _make_fill("2D6A9F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, h in enumerate(headers, 1):
        cell = ws_result.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    ws_result.row_dimensions[1].height = 30

    # データ行
    for row_idx, r in enumerate(results, 2):
        status = r.get("status", STATUS_OK)
        fill = _make_fill(STATUS_COLOR_MAP.get(status, COLOR_OK))
        border = _thin_border()

        # 出勤打刻情報
        att = r.get("matched_attendance")
        clock_in_str = ""
        clock_out_str = ""
        ref_source = ""
        if att:
            best = att[0]
            clock_in_str = _fmt_time(best.get("clock_in"))
            clock_out_str = _fmt_time(best.get("clock_out"))
            ref_source = "タイムカード"
        elif r.get("matched_contract"):
            c = r["matched_contract"]
            clock_in_str = _fmt_time(c.get("work_start"))
            clock_out_str = _fmt_time(c.get("work_end"))
            ref_source = "雇用契約書"

        alerts_str = " / ".join(r.get("alerts", []))
        status_label = STATUS_LABEL_MAP.get(status, status)

        row_data = [
            r.get("raw_row_index", row_idx - 1),
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            _fmt_date(r.get("date")),
            _fmt_time(r.get("start_time")),
            _fmt_time(r.get("end_time")),
            r.get("video_title", ""),
            r.get("progress", ""),
            status_label,
            alerts_str,
            clock_in_str,
            clock_out_str,
            ref_source,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws_result.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 10))
            if col_idx == 10:  # 判定詳細列は折り返し
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 人物不一致は太字にして強調
        if status == STATUS_PERSON_MISMATCH:
            for col_idx in range(1, len(headers) + 1):
                ws_result.cell(row=row_idx, column=col_idx).font = Font(bold=True, color="8B0000")

    # 列幅調整
    col_widths = [8, 12, 14, 12, 10, 10, 30, 8, 18, 50, 12, 12, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws_result.column_dimensions[get_column_letter(col_idx)].width = width

    # ウィンドウ枠固定（1行目）
    ws_result.freeze_panes = "A2"

    # オートフィルター
    ws_result.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ============================================================
    # シート2: アラート一覧（問題のある行のみ）
    # ============================================================
    ws_alert = wb.create_sheet("アラート一覧")
    alert_results = [r for r in results if r.get("alerts") and r.get("status") != STATUS_OK]

    alert_headers = ["優先度", "社員番号", "氏名", "日付", "視聴開始", "視聴終了",
                     "動画タイトル", "ステータス", "アラート詳細"]
    for col_idx, h in enumerate(alert_headers, 1):
        cell = ws_alert.cell(row=1, column=col_idx, value=h)
        cell.fill = _make_fill("C00000")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    priority_map = {
        STATUS_PERSON_MISMATCH: 1,
        STATUS_OUTSIDE_HOURS: 2,
        STATUS_CONTRACT_OUTSIDE: 2,
        STATUS_NO_ATTENDANCE: 3,
        STATUS_CONTRACT_ONLY: 4,
        STATUS_NO_TIME_INFO: 5,
    }

    alert_results_sorted = sorted(
        alert_results,
        key=lambda r: (priority_map.get(r.get("status"), 9), r.get("date") or "")
    )

    for row_idx, r in enumerate(alert_results_sorted, 2):
        status = r.get("status", "")
        fill = _make_fill(STATUS_COLOR_MAP.get(status, COLOR_OK))
        priority = priority_map.get(status, 9)
        priority_label = {1: "🚨 最高", 2: "⚠ 高", 3: "⚠ 中", 4: "ℹ 低", 5: "- 確認"}.get(priority, "-")

        row_data = [
            priority_label,
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            _fmt_date(r.get("date")),
            _fmt_time(r.get("start_time")),
            _fmt_time(r.get("end_time")),
            r.get("video_title", ""),
            STATUS_LABEL_MAP.get(status, status),
            " / ".join(r.get("alerts", [])),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_alert.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 9))

    col_widths_alert = [12, 12, 14, 12, 10, 10, 30, 18, 60]
    for col_idx, width in enumerate(col_widths_alert, 1):
        ws_alert.column_dimensions[get_column_letter(col_idx)].width = width
    ws_alert.freeze_panes = "A2"
    if alert_results_sorted:
        ws_alert.auto_filter.ref = f"A1:{get_column_letter(len(alert_headers))}1"

    # ============================================================
    # シート3: サマリー
    # ============================================================
    ws_summary = wb.create_sheet("サマリー")
    summary = generate_summary(results)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 30

    # タイトル
    title_cell = ws_summary.cell(row=1, column=1, value="動画ログ照合 サマリーレポート")
    title_cell.font = Font(bold=True, size=14, color="2D6A9F")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[1].height = 30

    gen_time = ws_summary.cell(row=2, column=1, value=f"生成日時: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}")
    gen_time.font = Font(size=9, color="888888")

    rows = [
        ("", "", ""),
        ("項目", "件数", "説明"),
        ("動画ログ総件数", summary["total"], ""),
        ("✅ 正常（勤務時間内）", summary["ok"], "タイムカードで確認済み"),
        ("ℹ 契約時間参照（問題なし）", summary["contract_only"], "タイムカードなし・契約時間内"),
        ("⚠ 勤務時間外視聴", summary["outside_hours"], "要確認"),
        ("⚠ 契約時間外視聴", summary["contract_outside"], "要確認"),
        ("⚠ 出勤記録なし", summary["no_attendance"], "タイムカード・契約書共になし"),
        ("🚨 人物不一致", summary["person_mismatch"], "緊急確認が必要"),
        ("- 時刻情報なし", summary["no_time_info"], "動画ログに時刻未記載"),
        ("", "", ""),
        ("アラートあり件数合計", summary["alert_count"], "「アラート一覧」シートを参照"),
    ]

    for row_offset, (label, value, desc) in enumerate(rows, 3):
        ws_summary.cell(row=row_offset, column=1, value=label)
        if value != "":
            ws_summary.cell(row=row_offset, column=2, value=value)
        ws_summary.cell(row=row_offset, column=3, value=desc)

        if label == "項目":
            for col in range(1, 4):
                c = ws_summary.cell(row=row_offset, column=col)
                c.fill = _make_fill("2D6A9F")
                c.font = Font(bold=True, color="FFFFFF")
                c.alignment = Alignment(horizontal="center")

        elif "🚨" in str(label) and value > 0:
            for col in range(1, 3):
                c = ws_summary.cell(row=row_offset, column=col)
                c.fill = _make_fill(COLOR_PERSON_MISMATCH)
                c.font = Font(bold=True)
        elif "⚠" in str(label) and value > 0:
            for col in range(1, 3):
                c = ws_summary.cell(row=row_offset, column=col)
                c.fill = _make_fill(COLOR_OUTSIDE_HOURS)
        elif "✅" in str(label):
            for col in range(1, 3):
                c = ws_summary.cell(row=row_offset, column=col)
                c.fill = _make_fill(COLOR_OK)

    # ============================================================
    # シート4: 凡例
    # ============================================================
    ws_legend = wb.create_sheet("凡例")
    ws_legend.column_dimensions["A"].width = 20
    ws_legend.column_dimensions["B"].width = 15
    ws_legend.column_dimensions["C"].width = 45

    legend_title = ws_legend.cell(row=1, column=1, value="カラー凡例")
    legend_title.font = Font(bold=True, size=12)

    legend_data = [
        (COLOR_OK, "OK", "勤務時間内に視聴（タイムカード確認済み）"),
        (COLOR_CONTRACT_ONLY, "契約時間参照", "タイムカードなし・雇用契約の所定時間内"),
        (COLOR_OUTSIDE_HOURS, "⚠ 勤務時間外", "タイムカードの出退勤時間外に視聴"),
        (COLOR_CONTRACT_OUTSIDE, "⚠ 契約時間外", "雇用契約の所定時間外に視聴"),
        (COLOR_NO_ATTENDANCE, "⚠ 出勤記録なし", "当日の出勤記録も契約書も見つからない"),
        (COLOR_PERSON_MISMATCH, "🚨 人物不一致", "動画ログと出勤記録の人物が一致しない"),
        (COLOR_NO_TIME_INFO, "時刻情報なし", "動画ログに視聴時刻が記録されていない"),
    ]

    for row_offset, (color, label, desc) in enumerate(legend_data, 3):
        c1 = ws_legend.cell(row=row_offset, column=1, value="　")
        c1.fill = _make_fill(color)
        c1.border = _thin_border()
        c2 = ws_legend.cell(row=row_offset, column=2, value=label)
        c2.font = Font(bold=True)
        ws_legend.cell(row=row_offset, column=3, value=desc)

    # シート順序を調整（サマリーを先頭に）
    wb.move_sheet("サマリー", offset=-wb.index(wb["サマリー"]))

    # 出力
    if output_path:
        wb.save(output_path)
        return b""
    else:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
